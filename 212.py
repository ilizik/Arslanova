import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np


class DataSet:
    def __init__(self, file_name, vacancy_name):
        self.file_name = file_name
        self.vacancy_name = vacancy_name

    @staticmethod
    def increment(dict, key, count):
        if key in dict:
            dict[key] += count
        else:
            dict[key] = count

    @staticmethod
    def medium(dictionary):
        new_dictionary = {}
        for key, values in dictionary.items():
            new_dictionary[key] = int(sum(values) / len(values))
        return new_dictionary

    def csv_reader(self):
        with open(self.file_name, mode='r', encoding='utf-8-sig') as file:
            reader = csv.reader(file)
            header = next(reader)
            for row in reader:
                if '' not in row and len(row) == len(header):
                    yield dict(zip(header, row))

    def get_statistic(self):
        salary = {}
        salary_vacancy = {}
        city = {}
        count_vacancies = 0

        for vacancy_dictionary in self.csv_reader():
            vacancy = Vacancy(vacancy_dictionary)
            self.increment(salary, vacancy.year, [vacancy.salary_average])
            if vacancy.name.find(self.vacancy_name) != -1:
                self.increment(salary_vacancy, vacancy.year, [vacancy.salary_average])
            self.increment(city, vacancy.area_name, [vacancy.salary_average])
            count_vacancies += 1

        vacancies_number = dict([(k, len(v)) for k, v in salary.items()])
        vacancies_number_name = dict([(k, len(v)) for k, v in salary_vacancy.items()])

        if not salary_vacancy:
            salary_vacancy = dict([(key, [0]) for key, value in salary.items()])
            vacancies_number_name = dict([(k, 0) for k, v in vacancies_number.items()])

        stats = self.medium(salary)
        stats_2 = self.medium(salary_vacancy)
        stats_3 = self.medium(city)
        stats_4 = {}

        for year, salaries in city.items():
            stats_4[year] = round(len(salaries) / count_vacancies, 4)
        stats_4 = list(filter(lambda a: a[-1] >= 0.01, [(k, v) for k, v in stats_4.items()]))
        stats_4.sort(key=lambda a: a[-1], reverse=True)
        stats_5 = stats_4.copy()
        stats_4 = dict(stats_4)
        stats_3 = list(filter(lambda a: a[0] in list(stats_4.keys()), [(k, v) for k, v in stats_3.items()]))
        stats_3.sort(key=lambda a: a[-1], reverse=True)
        stats_3 = dict(stats_3[:10])
        stats_5 = dict(stats_5[:10])

        return stats, vacancies_number, stats_2, vacancies_number_name, stats_3, stats_5


class Report:
    def __init__(self, vacancy_name, stats_1, stats_2, stats_3, stats_4, stats_5, stats_6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats_1 = stats_1
        self.stats_2 = stats_2
        self.stats_3 = stats_3
        self.stats_4 = stats_4
        self.stats_5 = stats_5
        self.stats_6 = stats_6

    def get_excel(self):
        first_ws = self.wb.active
        first_ws.title = 'Статистика по годам'
        first_ws.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий',
                         'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats_1.keys():
            first_ws.append([year, self.stats_1[year], self.stats_3[year], self.stats_2[year], self.stats_4[year]])

        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancy_name]]
        widths_col = []
        for row in data:
            for i, cell in enumerate(row):
                if len(widths_col) > i:
                    if len(cell) > widths_col[i]:
                        widths_col[i] = len(cell)
                else:
                    widths_col += [len(cell)]

        for i, widths_col in enumerate(widths_col, 1):
            first_ws.column_dimensions[get_column_letter(i)].width = widths_col + 2

        data = [['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий']]

        for (first_city, first_value), (second_city, second_value) in zip(self.stats_5.items(), self.stats_6.items()):
            data.append([first_city, first_value, '', second_city, second_value])

        second_ws = self.wb.create_sheet('Статистика по городам')
        for row in data:
            second_ws.append(row)

        widths_col = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(widths_col) > i:
                    if len(cell) > widths_col[i]:
                        widths_col[i] = len(cell)
                else:
                    widths_col += [len(cell)]

        for i, column_width in enumerate(widths_col, 1):
            second_ws.column_dimensions[get_column_letter(i)].width = column_width + 2

        font_bold = Font(bold=True)
        for col in 'ABCDE':
            first_ws[col + '1'].font = font_bold
            second_ws[col + '1'].font = font_bold

        for index, _ in enumerate(self.stats_5):
            second_ws['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                second_ws[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        for row, _ in enumerate(self.stats_1):
            for col in 'ABCDE':
                first_ws[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

    def get_png(self):
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(nrows=2, ncols=2)

        bar1 = ax1.bar(np.array(list(self.stats_1.keys())) - 0.4, self.stats_1.values(), width=0.4)
        bar2 = ax1.bar(np.array(list(self.stats_1.keys())), self.stats_3.values(), width=0.4)
        ax1.set_title('Уровень зарплат по годам', fontdict={'fontsize': 8})
        ax1.grid(axis='y')
        ax1.legend((bar1[0], bar2[0]), ('средняя з/п', 'з/п ' + self.vacancy_name.lower()), prop={'size': 8})
        ax1.set_xticks(np.array(list(self.stats_1.keys())) - 0.2, list(self.stats_1.keys()), rotation=90)
        ax1.xaxis.set_tick_params(labelsize=8)
        ax1.yaxis.set_tick_params(labelsize=8)

        ax2.set_title('Количество вакансий по годам', fontdict={'fontsize': 8})
        bar1 = ax2.bar(np.array(list(self.stats_2.keys())) - 0.4, self.stats_2.values(), width=0.4)
        bar2 = ax2.bar(np.array(list(self.stats_2.keys())), self.stats_4.values(), width=0.4)
        ax2.legend((bar1[0], bar2[0]), ('Количество вакансий', 'Количество вакансий\n' + self.vacancy_name.lower()),
                   prop={'size': 8})
        ax2.set_xticks(np.array(list(self.stats_2.keys())) - 0.2, list(self.stats_2.keys()), rotation=90)
        ax2.grid(axis='y')
        ax2.xaxis.set_tick_params(labelsize=8)
        ax2.yaxis.set_tick_params(labelsize=8)

        ax3.set_title('Уровень зарплат по городам', fontdict={'fontsize': 8})
        ax3.barh(list([str(a).replace(' ', '\n').replace('-', '-\n') for a in reversed(list(self.stats_5.keys()))]),
                 list(reversed(list(self.stats_5.values()))), color='blue', height=0.5, align='center')
        ax3.yaxis.set_tick_params(labelsize=6)
        ax3.xaxis.set_tick_params(labelsize=8)
        ax3.grid(axis='x')

        ax4.set_title('Доля вакансий по городам', fontdict={'fontsize': 8})
        other = 1 - sum([value for value in self.stats_6.values()])
        ax4.pie(list(self.stats_6.values()) + [other], labels=list(self.stats_6.keys()) + ['Другие'],
                textprops={'fontsize': 6})

        plt.tight_layout()
        plt.savefig('graph.png')

    def save(self, filename):
        self.wb.save(filename=filename)


class Vacancy:
    currency_to_rub = {
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KGS": 0.76,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055
    }

    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency_to_rub[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])


class InputConnect:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = DataSet(self.file_name, self.vacancy_name)
        stats_1, stats_2, stats_3, stats_4, stats_5, stats_6 = dataset.get_statistic()

        report = Report(self.vacancy_name, stats_1, stats_2, stats_3, stats_4, stats_5, stats_6)
        report.get_excel()
        report.save('report.xlsx')
        report.get_png()


InputConnect()